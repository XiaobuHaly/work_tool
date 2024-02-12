import os
from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook
from datetime import datetime

# 定义作业目录路径
directory = r"work_tool\data"

# 创建 Tkinter 应用
app = Tk()
app.title("作业查看工具")

# 设置窗口的最小大小
app.minsize(600, 300)

# 设置窗口的初始大小
app.geometry('600x650')

# 来源选项，0 表示文件的父目录，1 表示文件名
source_option = 1

# 判断一行是否在合并单元格中
def in_merged_cells(row, merged_cells):
    for m in merged_cells:
        if m.min_row <= row <= m.max_row:
            return True, f"{m.min_row}.{m.max_row}" # 返回是合并单元格及对应的日期范围
    return False, None                      # 是合并单元格及对应的日期范围

# 格式化日期范围
def format_date_range(date_range):
    min_date, max_date = date_range.split('.')  # 分割日期范围
    if min_date == max_date:                # 如果是单个日期，
        return min_date                     # 返回单个日期
    else:
        return date_range                   # 否则返回日期范围

# 判断日期内容是否在合并单元格
def in_merged_cells(row, merged_cells):
    for m in merged_cells:                  # 遍历合并单元格
        if m.min_row <= row <= m.max_row:   # 如果行在合并单元格中
            return True, m                  # 返回是合并单元格及对应的合并单元格对象
    return False, None                      # 是合并单元格及对应的合并单元格对象

# 读取特定日期的作业内容
def read_homework(date):
    homework_content = ""                               # 作业内容
    for root, _, files in os.walk(directory):           # 遍历目录
        for file in files:                              # 遍历文件
            if file.endswith(".xlsx"):                  # 确保文件是Excel文件
                file_path = os.path.join(root, file)    # 文件路径
                if source_option == 0:                  # 文件的父目录
                    source = os.path.basename(os.path.dirname(file_path))   # 文件的父目录
                elif source_option == 1:                # 文件名
                    source = os.path.splitext(file)[0]  # 文件名
                wb = load_workbook(file_path)           # 读取Excel文件
                sheet = wb.active                       # 获取工作表

                for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):    # 遍历行
                    try:                                
                        excel_date = row[0].strftime('%m.%d')   # 日期
                    except AttributeError:              # 日期为None
                        excel_date = str(row[0])        

                    if excel_date != date:              # 如果日期不等于指定日期
                        continue                        # 继续下一行

                    in_merge, date_range = in_merged_cells(i, sheet.merged_cells)   # 判断是否在合并单元格
                    if in_merge:                        # 如果在合并单元格
                        homework_items = []             # 作业内容
                        item_counter = 0                # 作业项计数
                        for j in range(date_range.min_row, date_range.max_row+1):   # 遍历合并单元格
                            homework_item = [f"{idx+1 + item_counter}、{cell.value}" \
                                             for idx, cell in enumerate(sheet[j][1:]) if cell.value]    # 作业内容
                            homework_items.extend(homework_item)    # 添加作业内容
                            item_counter += len(homework_item)      # 作业项计数
                    else:
                        homework_items = [f"{idx+1}、 {item}" \
                                          for idx, item in enumerate(row[1:]) if item]  # 作业内容

                    if homework_items:
                        homework_content += f"来源：{source}\n"                 # 添加来源
                        homework_content += "\n".join(homework_items) + "\n\n"  # 添加作业内容
    return homework_content                             # 返回作业内容


# 显示特定日期的作业内容
def show_homework():                                    # 显示特定日期的作业内容
    date = date_entry.get()                             # 赋值日期
    if not date:                                        # 如果日期为空
        messagebox.showerror("错误", "请输入日期！")     # 弹出错误提示   
        return                                          # 返回
    homework_text.delete(1.0, END)                      # 清空作业内容
    homework = read_homework(date)                      # 读取作业内容
    homework_text.insert(END, homework if homework else "当天无作业")   # 显示作业内容

# 创建日期输入框和按钮
date_label = Label(app, text="请输入日期(格式：#m.#d，例如：1.29、2.3):")   # 创建日期标签
date_label.grid(row=0, column=0, padx=10, pady=10, sticky=W+E)           # 设置位置
date_entry = Entry(app)                                                  # 创建日期输入框
date_entry.grid(row=0, column=1, padx=10, pady=10, sticky=W+E)           # 设置位置
date_entry.insert(0, f"{datetime.today().month}.{datetime.today().day}") # 设置默认日期
show_button = Button(app, text="查看作业", command=show_homework)         # 创建查看作业按钮
show_button.grid(row=0, column=2, padx=10, pady=10, sticky=W+E)          # 设置位置

# 创建作业显示文本框
homework_text = Text(app, width=50, height=10)                           # 创建作业显示文本框
homework_text.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky=W+E+N+S) # 设置位置

# 窗口大小适应
app.grid_rowconfigure(1, weight=1)              # 设置行的权重
app.grid_columnconfigure(1, weight=1)           # 设置列的权重

# 运行
app.mainloop()
